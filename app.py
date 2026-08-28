"""
DCF & Relative Valuation Studio -- single-file build.
Screener data + beta vs Nifty + DCF / residual income / DDM / comps.
Run:  pip install -r requirements.txt && gunicorn app:app   (or: python3 app.py)
"""
from __future__ import annotations

import io
import json
import math
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from typing import Any, Optional

import numpy as np
import pandas as pd
import requests
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter



# ============== screener ==============
"""
screener.py -- client + parser for screener.in (Indian equities).

Verified behaviour (probed 2026-08-28 from this sandbox):
  * https://www.screener.in/company/<TICKER>/consolidated/      -> 200, full HTML,
    12y P&L / balance sheet / cash-flow tables are NOT masked for logged-out users.
  * https://www.screener.in/api/company/<warehouseId>/peers/    -> 200, real peer table
    (S.No, Name, CMP, P/E, Mar Cap, Div Yld, NP Qtr, Qtr Profit Var, Sales Qtr,
     Qtr Sales Var, ROCE) plus a "Median: NN Co." row.
  * https://www.screener.in/api/company/<id>/quick_ratios/      -> login-gated (Register page).
  * https://www.screener.in/company/export/                     -> 404 (login-gated).
  * https://www.screener.in/api/company/<TICKER>/ratios/        -> 404.
So: peers endpoint works, export endpoint does not. We therefore scrape the HTML
tables and re-emit a screener-style .xlsx ourselves (see py) so the user
gets the export file that screener.in will not give us anonymously.
"""




BASE = "https://www.screener.in"
UA = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)

_SESSION = threading.local()
# warehouseId is what the peers API wants (e.g. RELIANCE -> 2726)
_ID_CACHE: dict[str, str] = {}


def _sess() -> requests.Session:
    s = getattr(_SESSION, "s", None)
    if s is None:
        s = requests.Session()
        s.headers.update(
            {
                "User-Agent": UA,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            }
        )
        _SESSION.s = s
    return s


# --------------------------------------------------------------------------- #
# small helpers
# --------------------------------------------------------------------------- #
def _num(v: Any) -> Optional[float]:
    """'1,23,456' / '18%' / '-2,188' / '' -> float | None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if pd.isna(f) else f
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("\u2013", "-").replace("\u2212", "-").replace("\xa0", " ")
    if s in {"-", "--", "NA", "N.A.", "nan", "None"}:
        return None
    pct = s.endswith("%")
    s = s.replace(",", "").replace("%", "").replace(" ", "").replace("\u20b9", "")
    if s in {"", "-"}:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return f


def _strip_tags(html: str) -> str:
    txt = re.sub(r"<[^>]+>", " ", html)
    txt = (txt.replace("&amp;", "&").replace("&nbsp;", " ")
              .replace("&lt;", "<").replace("&gt;", ">").replace("&#39;", "'"))
    return re.sub(r"\s+", " ", txt).strip()


def _section(html: str, anchor: str, stop_anchors: tuple[str, ...]) -> str:
    i = html.find(anchor)
    if i < 0:
        return ""
    end = len(html)
    for stop in stop_anchors:
        j = html.find(stop, i + len(anchor))
        if 0 < j < end:
            end = j
    return html[i:end]


def _label(cell_html: str) -> str:
    return _strip_tags(cell_html).replace("+", "").strip()


def _row_cells(row_html: str) -> list[str]:
    return re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_html, re.S | re.I)


def table_to_frame(section_html: str) -> Optional[pd.DataFrame]:
    """First <table> in a section -> DataFrame (rows = line items, cols = periods)."""
    t0 = section_html.find("<table")
    if t0 < 0:
        return None
    t1 = section_html.find("</table>", t0)
    raw = section_html[t0: t1 if t1 > 0 else None]
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", raw, re.S | re.I)
    if not rows:
        return None

    header_cells = _row_cells(rows[0])
    cols = [_strip_tags(c) for c in header_cells[1:]]
    data: dict[str, dict[str, Any]] = {}
    for r in rows[1:]:
        cells = _row_cells(r)
        if len(cells) < 2:
            continue
        lab = _label(cells[0])
        if not lab:
            continue
        vals = [_num(_strip_tags(c)) for c in cells[1:]]
        # pad to header width
        vals += [None] * (len(cols) - len(vals))
        data.setdefault(lab, dict(zip(cols, vals[: len(cols)])))

    if not data:
        return None
    df = pd.DataFrame.from_dict(data, orient="index")
    df.columns = list(cols)
    return df


def _find(df: Optional[pd.DataFrame], *needles: str) -> Optional[str]:
    """First index label containing any needle (needles tried in priority order)."""
    if df is None:
        return None
    low = {str(i): str(i).lower() for i in df.index}
    for n in needles:
        for orig, l in low.items():
            if n in l:
                return orig
    return None


# --------------------------------------------------------------------------- #
# top-of-page metric strip
# --------------------------------------------------------------------------- #
TOP_METRIC_KEYS = [
    "Market Cap", "Current Price", "High / Low", "Stock P/E", "Book Value",
    "Dividend Yield", "ROCE", "ROE", "Face Value",
]


def parse_top_ratios(html: str) -> dict[str, Optional[float]]:
    out: dict[str, Optional[float]] = {}
    block = _section(html, 'id="top-ratios"', ("</ul>",))
    for name, val in re.findall(
        r'<span class="name">\s*([^<]+?)\s*</span>\s*<span class="nowrap value">(.*?)</span>',
        block, re.S,
    ):
        key = re.sub(r"\s+", " ", name).strip()
        txt = _strip_tags(val)
        out[key] = _num(txt)
    # "High / Low" renders as "1,612 / 1,250" -> _num returns None; split manually
    hl = re.search(r'High / Low.*?number">([\d,\.]+)</span>\s*/\s*<span class="number">([\d,\.]+)',
                   html, re.S)
    if hl:
        out["52W High"] = _num(hl.group(1))
        out["52W Low"] = _num(hl.group(2))
    return out


def parse_sector(html: str) -> dict[str, str]:
    """Screener's 4-level industry breadcrumb inside the peers card."""
    seg = _section(html, 'id="peers"', ('id="quarters"',))
    labels = re.findall(r'title="(Broad Sector|Sector|Broad Industry|Industry)"', seg)
    names = [_strip_tags(m) for m in re.findall(
        r'<a href="/market/[^"]+"[^>]*>(.*?)</a>', seg, re.S)]
    d = dict(zip(labels, names))
    d["industry_url"] = ""
    m = re.search(r'<a href="(/market/[^"]+)"[^>]*title="Industry"', seg)
    if m:
        d["industry_url"] = BASE + m.group(1)
    return d


def parse_warehouse_id(html: str) -> Optional[str]:
    if not html:
        return None
    m = re.search(r'data-company-id="(\d+)"', html)
    if m:
        return m.group(1)
    m = re.search(r"/api/company/(\d+)/", html)
    if m:
        return m.group(1)
    m = re.search(r'data-url="/notebook/(\d+)/"', html)
    return m.group(1) if m else None


def parse_name(html: str) -> str:
    m = re.search(r"<title>\s*(.*?)\s*share price", html, re.S)
    if m:
        return _strip_tags(m.group(1))
    m = re.search(r"<h1[^>]*>\s*(.*?)\s*</h1>", html, re.S)
    return _strip_tags(m.group(1)) if m else ""


# --------------------------------------------------------------------------- #
# peers
# --------------------------------------------------------------------------- #
PEER_COLS = ["cmp", "pe", "mar_cap", "div_yld", "np_qtr", "np_qtr_var",
             "sales_qtr", "sales_qtr_var", "roce"]


def pick_peers(all_peers: list[dict], our_mcap: Optional[float],
                max_peers: int = 6, lo: float = 0.15, hi: float = 6.0) -> list[dict]:
    """Screener's peer list mixes holding cos / NBFCs of wildly different size into the
    same industry bucket. Prefer size-comparable names: |mcap/our_mcap| closest to 1,
    inside [lo, hi]; if that yields too few, relax the band."""
    if not all_peers:
        return []
    if not our_mcap or our_mcap <= 0:
        return all_peers[:max_peers]

    def ratio(p):
        return (p.get("mar_cap") or 0) / our_mcap

    banded = [p for p in all_peers if lo <= ratio(p) <= hi]
    if len(banded) < max(3, max_peers - 2):
        banded = [p for p in all_peers if 0.05 <= ratio(p) <= 20]
    if not banded:
        banded = all_peers
    banded.sort(key=lambda p: abs(ratio(p) - 1.0))
    return banded[:max_peers]


def fetch_peers(warehouse_id: str, referer: str) -> dict:
    """Returns {peers: [...], median: {...}} from screener's peer-comparison API."""
    s = _sess()
    s.headers["Referer"] = referer
    s.headers["X-Requested-With"] = "XMLHttpRequest"
    try:
        html = s.get(f"{BASE}/api/company/{warehouse_id}/peers/", timeout=25).text
    except requests.RequestException:
        return {"peers": [], "median": None}
    finally:
        s.headers.pop("X-Requested-With", None)

    peers: list[dict] = []
    median: dict | None = None
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I):
        cells = _row_cells(row)
        if len(cells) < 3:
            continue
        texts = [_strip_tags(c) for c in cells]
        link = re.search(r'<a[^>]+href="(/company/[^"]+)"', row)
        if not link:
            # header row (S.No.) or the trailing "Median: NN Co." row.
            # Median row layout: ["" , "Median: 41 Co.", CMP, P/E, ...]
            joined = " ".join(texts)
            if joined.lower().startswith(("s.no", "median")) or "median:" in joined.lower():
                m = re.search(r"median:\s*(\d+)\s*co", joined, re.I)
                nums = [_num(t) for t in texts[2:]]
                if len(nums) >= len(PEER_COLS) and any(v is not None for v in nums):
                    median = dict(zip(PEER_COLS, nums[: len(PEER_COLS)]))
                    median["n_companies"] = int(m.group(1)) if m else None
            continue
        vals = [_num(t) for t in texts[2:]]
        vals += [None] * (len(PEER_COLS) - len(vals))
        rec = dict(zip(PEER_COLS, vals[: len(PEER_COLS)]))
        url = link.group(1)
        parts = [p for p in url.rstrip("/").split("/") if p]
        rec["ticker"] = parts[-2] if len(parts) >= 2 and parts[-1].lower() == "consolidated" else parts[-1]
        am = re.search(r"<a[^>]+>(.*?)</a>", row, re.S)
        rec["name"] = _strip_tags(am.group(1)) if am else rec["ticker"]
        rec["url"] = BASE + url
        peers.append(rec)

    peers = [p for p in peers if p.get("mar_cap")]
    peers.sort(key=lambda p: p["mar_cap"] or 0, reverse=True)
    return {"peers": peers, "median": median}


def enrich_peer(ticker: str) -> dict:
    """Fetch a peer's own page -> book value, ROE, debt, EBITDA(TTM), cash."""
    d = fetch_company(ticker, with_peers=False)
    if not d:
        return {}
    return {
        "book_value": d["top"].get("Book Value"),
        "roe": d["top"].get("ROE"),
        "roce": d["top"].get("ROCE"),
        "pe": d["top"].get("Stock P/E"),
        "cmp": d["top"].get("Current Price"),
        "mar_cap": d["top"].get("Market Cap"),
        "div_yld": d["top"].get("Dividend Yield"),
        "debt": d["history"]["debt"],
        "cash": d["history"]["cash"],
        "ebitda": d["history"]["ebitda"],
        "eps": d["history"]["eps"],
        "revenue": d["history"]["revenue"],
        "net_profit": d["history"]["net_profit"],
    }


# --------------------------------------------------------------------------- #
# main entry
# --------------------------------------------------------------------------- #
def _get(path: str, retries: int = 3, backoff: float = 1.2) -> Optional[str]:
    """GET with retry. screener.in throttles bursts of requests with short-lived
    4xx/5xx responses, which otherwise show up as 'company not found'."""
    s = _sess()
    last = None
    for i in range(retries):
        try:
            r = s.get(BASE + path, timeout=25)
            last = r
            if r.status_code == 200 and 'id="top-ratios"' in r.text:
                return r.text
            if r.status_code in (429, 500, 502, 503, 504):
                time.sleep(backoff * (i + 1))
                continue
            if r.status_code == 404:
                # a genuine "no such ticker" is also 404 -- retry once, then give up
                if i == 0:
                    time.sleep(0.6)
                    continue
                return None
        except requests.RequestException:
            time.sleep(backoff * (i + 1))
    return None


def fetch_company(ticker: str, with_peers: bool = True, max_peers: int = 6) -> Optional[dict]:
    """Pull one company off screener.in and normalise into our internal model."""
    ticker = ticker.strip().upper()
    html = _get(f"/company/{ticker}/consolidated/")
    if html is None:
        html = _get(f"/company/{ticker}/")
        canonical = f"/company/{ticker}/"
    else:
        canonical = f"/company/{ticker}/consolidated/"
    if html is None:
        return None

    wid = parse_warehouse_id(html)
    if wid:
        _ID_CACHE[ticker] = wid

    pnl = table_to_frame(_section(html, 'id="profit-loss"', ('id="balance-sheet"',)))
    bs = table_to_frame(_section(html, 'id="balance-sheet"', ('id="cash-flow"',)))
    cf = table_to_frame(_section(html, 'id="cash-flow"', ('id="ratios"', 'id="shareholding"')))
    ratios = table_to_frame(_section(html, 'id="ratios"', ('id="shareholding"', 'id="documents"')))

    hist = build_history(pnl, bs, cf, ratios)

    out: dict[str, Any] = {
        "ticker": ticker,
        "warehouse_id": wid,
        "name": parse_name(html),
        "url": BASE + canonical,
        "top": parse_top_ratios(html),
        "sector": parse_sector(html),
        "history": hist,
        "peers": {"peers": [], "median": None},
        "tables": {
            "pnl": frame_out(pnl),
            "bs": frame_out(bs),
            "cf": frame_out(cf),
            "ratios": frame_out(ratios),
        },
    }

    if with_peers and wid:
        pr = fetch_peers(wid, out["url"])
        allp = [p for p in pr["peers"] if p.get("ticker") and p["ticker"].upper() != ticker]
        peers = pick_peers(allp, out["top"].get("Market Cap"), max_peers=max_peers)
        if peers:
            with ThreadPoolExecutor(max_workers=min(3, len(peers))) as ex:
                extra = list(ex.map(lambda p: enrich_peer(p["ticker"]), peers))
            for p, e in zip(peers, extra):
                p.update({k: v for k, v in e.items() if v is not None})
        out["peers"] = {"peers": peers, "median": pr["median"], "all_count": len(pr["peers"])}
    return out


def frame_out(df: Optional[pd.DataFrame]) -> dict:
    if df is None:
        return []
    cols = [str(c) for c in df.columns]
    rows = []
    for idx, row in df.iterrows():
        rows.append({"label": str(idx),
                     "values": [None if (v is None or pd.isna(v)) else float(v) for v in row.tolist()]})
    return {"columns": cols, "rows": rows}


# --------------------------------------------------------------------------- #
# history normalisation -> the numeric block the DCF consumes
# --------------------------------------------------------------------------- #
# Screener uses two P&L layouts: manufacturers/services ("Sales"/"Operating Profit")
# and NBFCs ("Revenue"/"Financing Profit"). Both must resolve.
LBL = {
    "revenue":       ("sales", "revenue"),
    "ebitda":        ("operating profit", "financing profit"),
    "margin":        ("opm", "financing margin"),
    "depreciation":  ("depreciation",),
    "pbt":           ("profit before tax",),
    "tax_pct":       ("tax %", "tax rate"),
    "net_profit":    ("net profit",),
    "eps":           ("eps",),
    "interest":      ("interest",),
    "other_income":  ("other income",),
    "div_payout":    ("dividend payout",),
    "fcf_row":       ("free cash flow",),
    "equity":        ("equity capital", "share capital"),
    "reserves":      ("reserves",),
    "debt":          ("borrowing", "debt"),
    "investments":   ("investment",),
    "other_assets":  ("other assets",),
    "minority":      ("minority",),
    "cfo":           ("operating activity",),
    "cfi":           ("investing activity",),
    "cff":           ("financing activity",),
    "ncf":           ("net cash flow",),
    "ccc":           ("cash conversion cycle",),
    "wc_days":       ("working capital days",),
    "roce":          ("roce",),
}


def build_history(pnl, bs, cf, ratios=None) -> dict:
    def series(df, key):
        lab = _find(df, *LBL[key])
        if lab is None:
            return {}
        return {c: v for c, v in df.loc[lab].to_dict().items()
                if v is not None and not pd.isna(v)}

    rev = series(pnl, "revenue")
    ebit = series(pnl, "ebitda")
    opm = series(pnl, "margin")
    dep = series(pnl, "depreciation")
    pbt = series(pnl, "pbt")
    npat = series(pnl, "net_profit")
    eps = series(pnl, "eps")
    interest = series(pnl, "interest")
    other_inc = series(pnl, "other_income")
    dpr = series(pnl, "div_payout")
    taxpct = series(pnl, "tax_pct")
    fcf_row = series(cf, "fcf_row")

    equity = series(bs, "equity")
    reserves = series(bs, "reserves")
    borrow = series(bs, "debt")
    invest = series(bs, "investments")
    oth_assets = series(bs, "other_assets")
    minority = series(bs, "minority")

    cfo = series(cf, "cfo")
    cfi = series(cf, "cfi")
    cff = series(cf, "cff")

    # screener's ratios card: Debtor / Inventory / Payable / CCC / Working Capital Days / ROCE
    ccc = series(ratios, "ccc")
    wc_days = series(ratios, "wc_days")
    roce_s = series(ratios, "roce")

    years = [y for y in rev if y != "TTM"]
    ttm_has = "TTM" in rev
    latest_y = "TTM" if ttm_has else (years[-1] if years else None)
    latest_bs = years[-1] if years else None

    def latest(d: dict):
        """Prefer TTM, else most recent FY -- some tables (cash flow) have no TTM."""
        if not d:
            return None
        if latest_y == "TTM" and d.get("TTM") is not None:
            return d["TTM"]
        for y in reversed(years):
            if d.get(y) is not None:
                return d[y]
        return None

    revenue = latest(rev)
    ebitda_ttm = latest(ebit)
    dep_ttm = latest(dep)
    npat_ttm = latest(npat)
    eps_ttm = latest(eps)
    ebit_ttm = None
    if ebitda_ttm is not None and dep_ttm is not None:
        ebit_ttm = ebitda_ttm - dep_ttm
    elif latest(pbt) is not None and latest(interest) is not None:
        ebit_ttm = latest(pbt) + latest(interest)

    debt = borrow.get(latest_bs) if latest_bs else None
    cash = invest.get(latest_bs) if latest_bs else None
    if cash is None and latest_bs:
        cash = oth_assets.get(latest_bs)

    cfi_vals = [v for v in (cfi.get(y) for y in years[-3:]) if v is not None]
    capex_proxy = (sum(abs(v) for v in cfi_vals) / len(cfi_vals)) if cfi_vals else None
    cfo_ttm = latest(cfo)
    fcf_ttm = latest(fcf_row)

    net = None
    if latest_bs and (equity.get(latest_bs) is not None or reserves.get(latest_bs)):
        net = (equity.get(latest_bs) or 0) + (reserves.get(latest_bs) or 0)

    payout = latest(dpr)
    if payout is None and npat_ttm and latest(cff) is not None:
        payout = None  # financing outflow is not a clean dividend proxy; leave to yfinance

    out = {
        "years": years,
        "has_ttm": ttm_has,
        "latest_period": latest_y,
        "latest_bs_period": latest_bs,
        "revenue": revenue,
        "ebitda": ebitda_ttm,
        "ebit": ebit_ttm,
        "depreciation": dep_ttm,
        "net_profit": npat_ttm,
        "eps": eps_ttm,
        "pbt": latest(pbt),
        "interest": latest(interest),
        "other_income": latest(other_inc),
        "opm_pct": latest(opm),
        "tax_pct": latest(taxpct),
        "dividend_payout_pct": payout,
        "debt": debt,
        "cash": cash,
        "equity_capital": equity.get(latest_bs) if latest_bs else None,
        "reserves": reserves.get(latest_bs) if latest_bs else None,
        "minority_interest": minority.get(latest_bs) if latest_bs else None,
        "cfo": cfo_ttm,
        "fcf_reported": fcf_ttm,
        "capex_proxy": capex_proxy,
        "ccc_days": latest(ccc),
        "wc_days": latest(wc_days),
        "roce_pct_reported": latest(roce_s),
        "series": {
            "revenue": rev, "ebitda": ebit, "opm": opm, "depreciation": dep,
            "net_profit": npat, "eps": eps, "pbt": pbt, "cfo": cfo, "cfi": cfi,
            "fcf": fcf_row,
        },
    }
    out["ebitda_margin_pct"] = (ebitda_ttm / revenue * 100.0) if revenue and ebitda_ttm else None
    out["net_worth"] = net
    out["roe_pct"] = (npat_ttm / net * 100.0) if (npat_ttm and net) else None
    out["net_debt"] = ((debt or 0) - (cash or 0)) if (debt is not None or cash is not None) else None
    return out


def warehouse_id_for(ticker: str) -> Optional[str]:
    t = ticker.strip().upper()
    if t in _ID_CACHE:
        return _ID_CACHE[t]
    s = _sess()
    try:
        r = s.get(f"{BASE}/company/{t}/consolidated/", timeout=20)
    except requests.RequestException:
        return None
    return parse_warehouse_id(r.text) if r.status_code == 200 else None

# ============== valuation ==============
"""
valuation.py -- the model. Everything in Rs crore unless a field says "_ps" (per share).

Methods
-------
1. Unlevered free-cash-flow DCF (WACC discounting, Gordon terminal value), with an
   EV/EBITDA exit-multiple cross-check.
2. Two-stage dividend discount model.
3. Trading comparables (EV/EBITDA, P/E, P/B) off screener.in peers.
Plus sensitivity grid, scenarios, and market-implied expectations.
"""



DEFAULTS: dict[str, Any] = {
    # --- capital structure / discount rate ---
    "risk_free_pct": 6.85,      # India 10Y gsec ~ this level; editable
    "erp_pct": 6.50,            # mature-market equity risk premium
    "crp_pct": 0.00,            # extra country risk premium for India
    "beta": 1.0,
    "beta_source": "user",
    "target_debt_pct": None,    # None -> use actual D/(D+E)
    "cost_of_debt_pct": None,   # None -> effective interest / debt
    "tax_rate_pct": None,       # None -> screener Tax %
    # --- explicit period ---
    "years": 10,
    "revenue_growth_pct": None, # None -> auto from history
    "growth_fade_pct": 1.0,     # growth falls this many pp each year
    "min_growth_pct": None,     # None -> terminal growth
    "ebit_margin_pct": None,    # None -> latest reported
    "target_margin_pct": None,  # None -> latest reported (flat)
    "margin_fade_years": 5,
    "capex_pct_revenue": None,  # None -> 3y avg |CFI| / revenue
    "dep_pct_revenue": None,    # None -> D&A / revenue
    "nwc_pct_revenue": None,    # None -> 0
    # --- terminal ---
    "terminal_growth_pct": 6.0,
    "terminal_method": "gordon",     # gordon | exit_multiple
    "exit_ev_ebitda": None,          # used when terminal_method == exit_multiple
    "discount_convention": "end",    # end | mid
    # --- ddm ---
    "payout_pct": None,
    "target_payout_pct": None,
    "ddm_growth_pct": None,
    # --- comps ---
    "peer_bands_pct": 25,
}


# --------------------------------------------------------------------------- #
def _pct(x):
    return None if x is None else x / 100.0


def _cagr(first, last, n):
    if not first or not last or first <= 0 or last <= 0 or n <= 0:
        return None
    return (last / first) ** (1.0 / n) - 1.0


def _blend(a, b):
    vals = [v for v in (a, b) if v is not None]
    return sum(vals) / len(vals) if vals else None


def cagr(series: dict, years: list[str], n: int) -> Optional[float]:
    ys = [y for y in years if series.get(y) is not None]
    if len(ys) < n + 1:
        n = max(1, len(ys) - 1)
    if len(ys) < 2:
        return None
    return _cagr(series[ys[-1 - n]], series[ys[-1]], n)


def auto_growth(hist: dict, tg: float) -> dict:
    """3y revenue CAGR, faded toward terminal growth so the model is not extrapolating
    a boom year forever."""
    c3 = cagr(hist["series"].get("revenue", {}), hist["years"], 3)
    c5 = cagr(hist["series"].get("revenue", {}), hist["years"], 5)
    base = _blend(c3, c5)
    if base is None:
        base = tg
    # cap: never start above 22% growth or below terminal growth
    base = max(tg, min(base, 0.22))
    return {"g1": base, "cagr3": c3, "cagr5": c5}


# --------------------------------------------------------------------------- #
def cost_of_capital(p: dict, hist: dict) -> dict:
    debt = hist.get("debt") or 0.0
    cash = hist.get("cash") or 0.0
    mcap = p.get("market_cap") or 0.0
    net_worth = hist.get("net_worth") or 0.0
    if mcap <= 0 and net_worth > 0:
        mcap = net_worth * 1.5  # crude fallback only

    w_d = p.get("target_debt_pct")
    if w_d is None:
        denom = debt + mcap
        w_d = (debt / denom) if denom > 0 else 0.0
    w_d = max(0.0, min(1.0, w_d))
    w_e = 1.0 - w_d

    ke = _pct(p["risk_free_pct"]) + p["beta"] * (_pct(p["erp_pct"]) + _pct(p["crp_pct"]))

    kd = p.get("cost_of_debt_pct")
    if kd is None:
        interest = hist.get("interest")
        if interest and debt and debt > 0:
            kd = min(0.18, max(0.04, interest / debt))
        else:
            kd = 0.085
    kd = _pct(kd) if kd > 1 else kd  # allow 8.5 or 0.085

    t = p.get("tax_rate_pct")
    if t is None:
        t = hist.get("tax_pct")
    if t is None or t <= 0 or t > 45:
        t = 25.0
    if t > 1:
        t = t / 100.0

    wacc = w_e * ke + w_d * kd * (1.0 - t)
    return {
        "cost_of_equity_pct": ke * 100,
        "cost_of_debt_pct": kd * 100,
        "w_d": w_d, "w_e": w_e,
        "tax_rate_pct": t * 100,
        "wacc_pct": wacc * 100,
    }


# --------------------------------------------------------------------------- #
def run_dcf(p: dict, hist: dict, cc: dict) -> dict:
    rev0 = hist.get("revenue")
    if not rev0:
        raise ValueError("No revenue in the source data - a DCF needs at least the latest "
                         "revenue figure.")
    if not hist.get("ebitda") and not hist.get("ebit"):
        raise ValueError("No operating profit / EBITDA in the source data. Add an "
                         "'Operating Profit' row to the imported statements.")
    tg = _pct(p["terminal_growth_pct"])
    g0 = _pct(p["revenue_growth_pct"]) if p["revenue_growth_pct"] is not None \
        else auto_growth(hist, tg)["g1"]
    fade = _pct(p["growth_fade_pct"])
    n = int(p["years"])

    m0 = _pct(p["ebit_margin_pct"]) if p["ebit_margin_pct"] is not None \
        else _pct(hist.get("ebitda_margin_pct"))
    mtgt = _pct(p["target_margin_pct"]) if p["target_margin_pct"] is not None else m0
    fade_yrs = max(1, int(p["margin_fade_years"]))

    cpx = _pct(p["capex_pct_revenue"]) if p["capex_pct_revenue"] is not None else (
        (hist.get("capex_proxy") or 0.0) / rev0 if rev0 else 0.10)
    dpr = _pct(p["dep_pct_revenue"]) if p["dep_pct_revenue"] is not None else (
        (hist.get("depreciation") or 0.0) / rev0 if rev0 else 0.05)
    nwc = _pct(p["nwc_pct_revenue"] or 0.0)

    wacc = _pct(cc["wacc_pct"])
    tax = _pct(cc["tax_rate_pct"])
    mid = p["discount_convention"] == "mid"

    rows, pv_explicit = [], 0.0
    prev_rev = rev0
    for i in range(1, n + 1):
        g = max(tg, g0 - fade * (i - 1))
        m = m0 + (mtgt - m0) * min(1.0, i / fade_yrs)
        rev = prev_rev * (1 + g)
        ebitda = rev * m
        dep = rev * dpr
        ebit = ebitda - dep
        capex = rev * cpx
        d_nwc = (rev - prev_rev) * nwc
        fcf = ebit * (1 - tax) + dep - capex - d_nwc
        disc = 1 / (1 + wacc) ** (i - 0.5 if mid else i)
        pv = fcf * disc
        pv_explicit += pv
        rows.append({
            "year": i, "growth_pct": g * 100, "margin_pct": m * 100,
            "revenue": rev, "ebitda": ebitda, "dep": dep, "ebit": ebit,
            "nopat": ebit * (1 - tax), "capex": capex, "d_nwc": d_nwc,
            "fcf": fcf, "pv_fcf": pv, "cum_pv": pv_explicit,
        })
        prev_rev = rev

    last = rows[-1]
    if p["terminal_method"] == "exit_multiple" and p.get("exit_ev_ebitda"):
        tv = last["ebitda"] * p["exit_ev_ebitda"]
        tv_note = f"EV/EBITDA {p['exit_ev_ebitda']:.1f}x on year-{n} EBITDA"
    else:
        # steady state: capex == depreciation, so UFCF = NOPAT - dNWC
        d_nwc_t = last["revenue"] * tg * nwc
        fcf_t1 = last["ebit"] * (1 - tax) - d_nwc_t
        if wacc <= tg:
            raise ValueError("WACC must exceed terminal growth rate")
        tv = fcf_t1 / (wacc - tg)
        tv_note = f"Gordon growth at {tg*100:.1f}% on normalised year-{n+1} FCF (capex = D&A)"
    pv_tv = tv / (1 + wacc) ** (n - 0.5 if mid else n)

    ev = pv_explicit + pv_tv
    net_debt = (hist.get("debt") or 0.0) - (hist.get("cash") or 0.0)
    minority = hist.get("minority_interest") or 0.0
    equity = ev - net_debt - minority
    shares = p.get("shares")
    vps = equity / shares if shares else None
    px = p.get("price")
    upside = ((vps / px - 1) * 100) if (vps and px) else None

    return {
        "rows": rows,
        "terminal_value": tv, "pv_terminal": pv_tv, "tv_note": tv_note,
        "pv_explicit": pv_explicit, "ev": ev, "net_debt": net_debt,
        "minority": minority, "equity_value": equity,
        "value_per_share": vps, "upside_pct": upside,
        "tv_share_of_ev_pct": (pv_tv / ev * 100) if ev else None,
        "implied_ev_ebitda_t0": (ev / hist["ebitda"]) if hist.get("ebitda") else None,
        "implied_pe_t0": (equity / hist["net_profit"]) if hist.get("net_profit") else None,
        "assumed": {
            "g1_pct": g0 * 100, "fade_pct": fade * 100, "m0_pct": m0 * 100,
            "mtgt_pct": mtgt * 100, "capex_pct": cpx * 100, "dep_pct": dpr * 100,
            "nwc_pct": nwc * 100, "terminal_growth_pct": tg * 100,
        },
    }


def sensitivity(p, hist, cc, dcf):
    """Value per share across a WACC x terminal-growth grid. WACC is moved by scaling
    beta so the rest of the capital structure stays intact."""
    base_w = cc["wacc_pct"]
    base_g = p["terminal_growth_pct"]
    ws = [round(base_w + d, 2) for d in (-1.5, -1.0, -0.5, 0.0, 0.5, 1.0, 1.5)]
    gs = [round(base_g + d, 2) for d in (-2.0, -1.0, 0.0, 1.0, 2.0)]

    def beta_for(target_wacc: float) -> Optional[float]:
        spread = _pct(p["erp_pct"]) + _pct(p["crp_pct"])
        if not spread or cc["w_e"] <= 0:
            return None
        kd = _pct(cc["cost_of_debt_pct"])
        t = _pct(cc["tax_rate_pct"])
        ke_needed = (target_wacc / 100.0 - cc["w_d"] * kd * (1 - t)) / cc["w_e"]
        beta = (ke_needed - _pct(p["risk_free_pct"])) / spread
        return beta if 0.0 <= beta <= 5.0 else None

    grid, valid = [], []
    for w in ws:
        row = []
        b = beta_for(w)
        for g in gs:
            v = None
            if b is not None and w / 100.0 > g / 100.0:
                try:
                    pp = dict(p)
                    pp["beta"] = b
                    pp["terminal_growth_pct"] = g
                    d2 = run_dcf(pp, hist, cost_of_capital(pp, hist))
                    v = d2["value_per_share"]
                except Exception:
                    v = None
            row.append(v)
            if v:
                valid.append(v)
        grid.append({"wacc": w, "values": row})
    return {"waccs": ws, "growths": gs, "grid": grid,
            "min": min(valid) if valid else None, "max": max(valid) if valid else None}


def scenarios(p, hist, cc):
    out = []
    for name, dg, dm, dw in (("Bear", -3.0, -2.0, 1.5), ("Base", 0.0, 0.0, 0.0),
                             ("Bull", 3.0, 2.0, -1.0)):
        pp = dict(p)
        g0 = p["revenue_growth_pct"] if p["revenue_growth_pct"] is not None else \
            auto_growth(hist, _pct(p["terminal_growth_pct"]))["g1"] * 100
        pp["revenue_growth_pct"] = max(pp["terminal_growth_pct"], g0 + dg)
        if pp["ebit_margin_pct"] is not None or hist.get("ebitda_margin_pct"):
            m0 = pp["ebit_margin_pct"] if pp["ebit_margin_pct"] is not None \
                else hist["ebitda_margin_pct"]
            pp["ebit_margin_pct"] = max(1.0, m0 + dm)
            if pp["target_margin_pct"] is None:
                pp["target_margin_pct"] = pp["ebit_margin_pct"]
            else:
                pp["target_margin_pct"] = max(1.0, pp["target_margin_pct"] + dm)
        pp["risk_free_pct"] = p["risk_free_pct"] + dw
        try:
            c2 = cost_of_capital(pp, hist)
            d2 = run_dcf(pp, hist, c2)
            out.append({"name": name, "value_per_share": d2["value_per_share"],
                        "upside_pct": d2["upside_pct"], "wacc_pct": c2["wacc_pct"],
                        "growth_pct": pp["revenue_growth_pct"],
                        "margin_pct": pp["ebit_margin_pct"], "ev": d2["ev"]})
        except Exception as e:
            out.append({"name": name, "error": str(e)})
    return out


# --------------------------------------------------------------------------- #
def ddm(p, hist, cc):
    """Two-stage dividend discount model. Needs DPS = payout x EPS."""
    ke = _pct(cc["cost_of_equity_pct"])
    eps = hist.get("eps")
    payout = p.get("payout_pct")
    if payout is None:
        payout = p.get("dividend_payout_pct")
    if payout is None:
        payout = hist.get("dividend_payout_pct")
    if payout is None:
        payout = 25.0
    tgt_payout = p.get("target_payout_pct") or payout
    g_ddm = p.get("ddm_growth_pct")
    if g_ddm is None:
        g_ddm = p["revenue_growth_pct"] if p["revenue_growth_pct"] is not None else 8.0
    g_ddm = min(g_ddm, ke * 100 - 1.0)
    tg = _pct(p["terminal_growth_pct"])
    n = int(p["years"])
    eps0 = eps or (hist.get("net_profit") / p["shares"] if p.get("shares") and hist.get("net_profit") else None)
    if not eps0:
        return {"error": "no EPS available for DDM"}

    rows, pv = [], 0.0
    e = eps0
    for i in range(1, n + 1):
        g = max(tg * 100, g_ddm - ((g_ddm - tg * 100) / max(n, 1)) * (i - 1))
        e *= (1 + g / 100)
        po = payout + (tgt_payout - payout) * min(1.0, i / max(1, n // 2))
        d = e * po / 100
        f = 1 / (1 + ke) ** i
        pv += d * f
        rows.append({"year": i, "eps": e, "payout_pct": po, "dps": d, "pv": d * f})
    d_t1 = rows[-1]["eps"] * tgt_payout / 100
    if ke <= tg:
        return {"error": "cost of equity must exceed terminal growth"}
    tv = d_t1 / (ke - tg)
    pv_tv = tv / (1 + ke) ** n
    return {"rows": rows, "stage1_pv": pv, "terminal_value": tv, "pv_terminal": pv_tv,
            "value_per_share": pv + pv_tv, "ke_pct": ke * 100,
            "terminal_dps": d_t1, "payout_pct": payout, "target_payout_pct": tgt_payout,
            "growth_pct": g_ddm}


# --------------------------------------------------------------------------- #
def _median(vals):
    v = sorted(x for x in vals if x is not None and not (isinstance(x, float) and math.isnan(x)))
    if not v:
        return None
    n = len(v)
    return v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2


def relative(p, hist, cc, peers, dcf):
    """Trading comparables. Multiples come from enriched screener peers."""
    band = p.get("peer_bands_pct") or 25
    recs = []
    for q in peers or []:
        mc = q.get("mar_cap")
        nd = ((q.get("debt") or 0) - (q.get("cash") or 0))
        ev = (mc + nd) if mc is not None else None
        ebitda = q.get("ebitda")
        bv = q.get("book_value")
        sh = (mc / q["cmp"]) if (mc and q.get("cmp")) else None
        nav = (bv * sh) if (bv is not None and sh) else None
        recs.append({
            "name": q.get("name"), "ticker": q.get("ticker"), "cmp": q.get("cmp"),
            "mar_cap": mc, "pe": q.get("pe"), "roe": q.get("roe"), "roce": q.get("roce"),
            "div_yld": q.get("div_yld"), "ev": ev, "ebitda": ebitda,
            "ev_ebitda": (ev / ebitda) if (ev and ebitda and ebitda > 0) else None,
            "pb": (q["cmp"] / bv) if (q.get("cmp") and bv) else None,
            "book_value": bv, "net_worth": nav,
        })
    med = {
        "pe": _median([r["pe"] for r in recs]),
        "ev_ebitda": _median([r["ev_ebitda"] for r in recs]),
        "pb": _median([r["pb"] for r in recs]),
        "div_yld": _median([r["div_yld"] for r in recs]),
        "roe": _median([r["roe"] for r in recs]),
        "roce": _median([r["roce"] for r in recs]),
        "mar_cap": _median([r["mar_cap"] for r in recs]),
    }
    lo_m, hi_m = 1 - band / 100, 1 + band / 100

    out = []
    # EV/EBITDA -> EV -> equity -> per share
    if med["ev_ebitda"] and hist.get("ebitda") and p.get("shares"):
        for tag, m in (("Low", med["ev_ebitda"] * lo_m), ("Median", med["ev_ebitda"]),
                       ("High", med["ev_ebitda"] * hi_m)):
            ev = m * hist["ebitda"]
            eq = ev - ((hist.get("debt") or 0) - (hist.get("cash") or 0))
            out.append({"metric": f"EV/EBITDA {m:.1f}x", "bucket": tag,
                        "multiple": m, "value": eq,
                        "per_share": eq / p["shares"] if p.get("shares") else None})
    if med["pe"] and hist.get("eps"):
        for tag, m in (("Low", med["pe"] * lo_m), ("Median", med["pe"]),
                       ("High", med["pe"] * hi_m)):
            eq = m * hist["net_profit"] if hist.get("net_profit") else m * hist["eps"] * p.get("shares", 1)
            out.append({"metric": f"P/E {m:.1f}x", "bucket": tag, "multiple": m,
                        "value": eq, "per_share": m * hist["eps"]})
    if med["pb"] and hist.get("net_worth") and p.get("shares"):
        bvps = hist["net_worth"] / p["shares"]
        for tag, m in (("Low", med["pb"] * lo_m), ("Median", med["pb"]),
                       ("High", med["pb"] * hi_m)):
            out.append({"metric": f"P/B {m:.2f}x", "bucket": tag, "multiple": m,
                        "value": m * hist["net_worth"], "per_share": m * bvps})

    grouped = {}
    for r in out:
        grouped.setdefault(r["metric"].split(" ")[0], []).append(r)
    return {"peers": recs, "median": med, "implied": out, "grouped": grouped,
            "our": {
                "pe": p.get("pe"), "pb": (p.get("price") / (hist["net_worth"] / p["shares"]))
                if (hist.get("net_worth") and p.get("shares")) else None,
                "ev_ebitda": ((p.get("market_cap") or 0) + (hist.get("debt") or 0)
                              - (hist.get("cash") or 0)) / hist["ebitda"]
                if hist.get("ebitda") else None,
                "div_yld": p.get("dividend_yield_pct"),
                "roe": hist.get("roe_pct"), "roce": p.get("roce"),
            }}


# --------------------------------------------------------------------------- #
def implied_expectations(p, hist, cc, dcf):
    """What growth does today's price embed? Solve the stage-1 growth that makes the
    DCF equal the market price (terminal growth held constant)."""
    px = p.get("price")
    if not px or not hist.get("revenue"):
        return None

    def value_at(g1_pct):
        pp = dict(p)
        pp["revenue_growth_pct"] = g1_pct
        try:
            return run_dcf(pp, hist, cc)["value_per_share"]
        except Exception:
            return None

    lo, hi = p["terminal_growth_pct"] - 5, 40.0
    vlo, vhi = value_at(lo), value_at(hi)
    if vlo is None or vhi is None or not (vlo < px < vhi):
        return {"error": "market price outside the solvable growth range",
                "low_growth_pct": lo, "high_growth_pct": hi,
                "low_value": vlo, "high_value": vhi}
    for _ in range(60):
        mid = (lo + hi) / 2
        v = value_at(mid)
        if v is None:
            break
        if v < px:
            lo = mid
        else:
            hi = mid
    return {"implied_growth_pct": (lo + hi) / 2,
            "terminal_growth_pct": p["terminal_growth_pct"],
            "wacc_pct": cc["wacc_pct"]}


def football_field(dcf, ddmres, rel, sens, scens=None):
    """Valuation ranges by method, for the football-field chart."""
    bars = []
    if dcf and dcf.get("value_per_share"):
        v = dcf["value_per_share"]
        svals = [s["value_per_share"] for s in (scens or [])
                 if s.get("value_per_share") is not None]
        bars.append({"method": "DCF - unlevered FCF", "low": min(svals) if svals else v * 0.7,
                     "mid": v, "high": max(svals) if svals else v * 1.3})
    if ddmres and ddmres.get("value_per_share"):
        v = ddmres["value_per_share"]
        bars.append({"method": "DDM - two stage", "low": v * 0.75, "mid": v, "high": v * 1.25})
    if rel and rel.get("implied"):
        for metric, rows in rel["grouped"].items():
            vals = [r["per_share"] for r in rows if r.get("per_share")]
            if vals:
                bars.append({"method": f"Relative - {metric}", "low": min(vals),
                             "mid": sorted(vals)[len(vals) // 2], "high": max(vals)})
    if sens and sens.get("min") and sens.get("max") and dcf:
        bars.append({"method": "DCF - WACC/growth grid", "low": sens["min"],
                     "mid": dcf["value_per_share"], "high": sens["max"]})
    return bars


# --------------------------------------------------------------------------- #
def warnings(p, hist, cc, dcf, rel, financial=False):
    w = []
    if financial:
        w.append("This is a bank / NBFC / insurer. Its borrowings are deposits and customer "
                 "funds, not capital structure, so unlevered free cash flow is not meaningful "
                 "and the DCF below should be ignored. Use the residual income value instead, "
                 "which is shown at the top of the football field.")
    if dcf:
        if dcf.get("tv_share_of_ev_pct") and dcf["tv_share_of_ev_pct"] > 90:
            w.append(f"Terminal value is {dcf['tv_share_of_ev_pct']:.0f}% of enterprise value - "
                     "the answer is mostly a perpetuity assumption, treat it with caution.")
        if dcf.get("value_per_share") and dcf["value_per_share"] < 0:
            w.append("DCF produces a negative equity value: free cash flow is too weak to cover "
                     "the cost of capital plus net debt.")
    if cc["wacc_pct"] <= p["terminal_growth_pct"]:
        w.append("WACC is at or below terminal growth - the Gordon formula breaks down.")
    if dcf and dcf.get("pv_explicit") is not None and dcf["pv_explicit"] < 0:
        w.append("The explicit forecast period generates negative free cash flow, so almost the "
                 "entire value rests on the terminal value. Check the working-capital and capex "
                 "assumptions before trusting this number.")
    nwc = (p.get("nwc_pct_revenue") or 0)
    if nwc >= 15:
        w.append(f"Working capital is assumed at {nwc:.0f}% of revenue, which is high. Inventory-heavy "
                 "businesses (jewellery, retail) show long cash-conversion cycles that are partly "
                 "financed by customer advances rather than by cash -- lower this in Assumptions "
                 "if that applies.")
    if (hist.get("net_debt") or 0) > 0 and hist.get("ebitda") and \
            hist["net_debt"] / hist["ebitda"] > 5:
        w.append(f"Net debt / EBITDA is {hist['net_debt']/hist['ebitda']:.1f}x - leverage risk "
                 "dominates the equity bridge.")
    if rel and rel.get("peers") and len(rel["peers"]) < 3:
        w.append(f"Only {len(rel['peers'])} usable comparables - the relative valuation is thin. "
                 "Add tickers in the Comps tab.")
    if hist.get("ebit") and hist["ebit"] < 0:
        w.append("Latest EBIT is negative; the DCF is running on a loss-making base.")
    return w


# --------------------------------------------------------------------------- #
# financials
# --------------------------------------------------------------------------- #
FINANCIAL_HINTS = (
    "bank", "financial", "nbfc", "insurance", "housing finance",
    "asset management", "stockbroking", "credit rating", "investment trust",
)


def is_financial(sector: dict, hist: dict) -> bool:
    """Banks, NBFCs and insurers fund themselves with deposits/borrowings, so
    'debt' is raw material rather than capital structure and EBIT is meaningless.
    Detect them so we can switch models instead of returning a nonsense number."""
    blob = " ".join(str(sector.get(k) or "").lower()
                    for k in ("Industry", "Broad Industry", "Sector", "Broad Sector"))
    if any(h in blob for h in FINANCIAL_HINTS):
        return True
    rev, debt = hist.get("revenue") or 0, hist.get("debt") or 0
    return bool(rev and debt > 3 * rev)


def residual_income(p, hist, cc):
    """Residual income (Edwards-Bell-Ohlson): equity value = book value plus the
    present value of earnings in excess of the equity charge. This is the right
    lens for banks and NBFCs, where unlevered FCF has no meaning."""
    ke = _pct(cc["cost_of_equity_pct"])
    nw = hist.get("net_worth")
    npat = hist.get("net_profit")
    shares = p.get("shares")
    if not nw or not npat or not shares:
        return {"error": "Residual income needs net worth, net profit and share count."}

    g0 = _pct(p["revenue_growth_pct"]) if p["revenue_growth_pct"] is not None else 0.12
    fade = _pct(p["growth_fade_pct"])
    tg = _pct(p["terminal_growth_pct"])
    n = int(p["years"])

    roe = npat / nw
    bv, rows, pv = nw, [], 0.0
    for i in range(1, n + 1):
        g = max(tg, g0 - fade * (i - 1))
        ri = bv * (roe - ke)
        f = 1 / (1 + ke) ** i
        pv += ri * f
        rows.append({"year": i, "growth_pct": g * 100, "book_value": bv, "roe_pct": roe * 100,
                     "equity_charge": bv * ke, "net_income": bv * roe,
                     "residual_income": ri, "pv": ri * f})
        bv *= (1 + g)

    # terminal: residual income persists at the faded return spread
    ri_t = rows[-1]["residual_income"] * (1 + tg)
    if ke <= tg:
        return {"error": "cost of equity must exceed terminal growth"}
    tv = ri_t / (ke - tg)
    pv_tv = tv / (1 + ke) ** n
    eq = nw + pv + pv_tv
    vps = eq / shares
    px = p.get("price")
    return {"rows": rows, "book_value": nw, "roe_pct": roe * 100, "ke_pct": ke * 100,
            "pv_residual": pv, "terminal_value": tv, "pv_terminal": pv_tv,
            "equity_value": eq, "value_per_share": vps,
            "upside_pct": ((vps / px - 1) * 100) if px else None,
            "implied_pb": (vps / (nw / shares)) if nw else None}

# ============== excel_io ==============
"""
py -- screener.in-style Excel workbook builder.

screener.in's own /company/export/ endpoint returns 404 for logged-out requests
(verified), so we re-emit the same shape from the HTML we already parsed: one sheet
with the annual P&L, balance sheet, cash flow and ratio blocks, exactly like their
download. That file can then be re-uploaded through /api/import for offline work.
"""




BLOCKS = [
    ("Consolidated Profit & Loss", "pnl"),
    ("Consolidated Balance Sheet", "bs"),
    ("Consolidated Cash Flow", "cf"),
    ("Ratios", "ratios"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
SUB_FILL = PatternFill("solid", fgColor="E5E7EB")
HEAD_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(bottom=THIN)


def build(tables: dict, name: str = "Company", ticker: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.freeze_panes = "B1"

    r = 1
    ws.cell(r, 1, name or ticker or "Company").font = Font(bold=True, size=13)
    r += 1
    ws.cell(r, 1, f"All values in Rs crore unless noted. Source: screener.in (parsed {ticker})")
    r += 2

    for title, key in BLOCKS:
        t = tables.get(key)
        if not t or not t.get("rows"):
            continue
        ws.cell(r, 1, title).font = BOLD
        ws.cell(r, 1).fill = SUB_FILL
        r += 1
        ws.cell(r, 1, "").fill = HEAD_FILL
        for j, col in enumerate(t["columns"], start=2):
            c = ws.cell(r, j, col)
            c.font = HEAD_FONT
            c.fill = HEAD_FILL
            c.alignment = Alignment(horizontal="right")
        r += 1
        for row in t["rows"]:
            lab = ws.cell(r, 1, row["label"])
            lab.border = BOX
            for j, v in enumerate(row["values"], start=2):
                c = ws.cell(r, j, v)
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
                c.border = BOX
            if row["label"].lower() in {"operating profit", "net profit", "total assets",
                                        "total liabilities", "net cash flow"}:
                for j in range(1, len(t["columns"]) + 2):
                    ws.cell(r, j).font = BOLD
            r += 1
        r += 2

    ws.column_dimensions["A"].width = 34
    for j in range(2, 22):
        ws.column_dimensions[get_column_letter(j)].width = 13

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ============== datafeed ==============
"""
py -- assembles one company model from whichever sources are available.

Priority: screener.in (financials, ratios, peers) -> yfinance (beta, payout, shares
cross-check). screener.in does not publish beta, so beta is always computed from
price history vs the Nifty unless the user types one.
"""




BENCH = "^NSEI"
_beta_cache: dict[str, tuple[float, float]] = {}   # ticker -> (ts, beta)
BETA_TTL = 6 * 3600


# --------------------------------------------------------------------------- #
def compute_beta(ticker: str, years: int = 5) -> Optional[float]:
    """OLS beta of weekly returns vs Nifty. Falls back to yfinance's own beta field."""
    now = time.time()
    hit = _beta_cache.get(ticker)
    if hit and now - hit[0] < BETA_TTL:
        return hit[1]
    try:
        import yfinance as yf
    except ImportError:
        return None
    try:
        px = yf.download([f"{ticker}.NS", BENCH], period=f"{years}y", interval="1wk",
                         auto_adjust=True, progress=False, threads=False)["Close"].dropna()
        if px.shape[0] > 40:
            r = px.pct_change().dropna()
            a = r.iloc[:, 0].to_numpy(dtype=float)
            b = r.iloc[:, 1].to_numpy(dtype=float)
            cov = ((a - a.mean()) * (b - b.mean())).mean()
            var = ((b - b.mean()) ** 2).mean()
            if var > 0:
                beta = float(cov / var)
                if 0.05 <= beta <= 3.5:
                    _beta_cache[ticker] = (now, round(beta, 3))
                    return round(beta, 3)
    except Exception:
        pass
    # fallback: Yahoo's own (often poorly estimated) beta field
    try:
        import yfinance as yf
        info = yf.Ticker(f"{ticker}.NS").info
        b = info.get("beta")
        if isinstance(b, (int, float)) and 0.05 <= b <= 3.5:
            _beta_cache[ticker] = (now, float(b))
            return float(b)
    except Exception:
        pass
    return None


def yfinance_extras(ticker: str) -> dict:
    """Payout, EPS and shares cross-check. Absolute rupees -> converted by caller."""
    try:
        import yfinance as yf
        info = yf.Ticker(f"{ticker}.NS").info or {}
    except Exception:
        return {}
    out = {}
    for src, dst in [("sharesOutstanding", "shares"), ("trailingPE", "pe"),
                     ("priceToBook", "pb"), ("dividendYield", "dividend_yield_pct"),
                     ("beta", "yahoo_beta"), ("trailingEps", "eps"),
                     ("currentPrice", "price"), ("marketCap", "market_cap"),
                     ("dividendRate", "dps"), ("bookValue", "book_value")]:
        v = info.get(src)
        if isinstance(v, (int, float)) and v == v:
            out[dst] = float(v)
    if out.get("eps") and out.get("dps"):
        out["payout_pct"] = out["dps"] / out["eps"] * 100.0
    elif out.get("eps") and out.get("dividend_yield_pct") and out.get("price"):
        out["payout_pct"] = (out["price"] * out["dividend_yield_pct"] / 100.0) / out["eps"] * 100.0
    return out


# --------------------------------------------------------------------------- #
def build(ticker: str, custom_peers: Optional[list[str]] = None,
          peer_mode: str = "curated", max_peers: int = 6) -> dict:
    """Returns {company, inputs, provenance, peers, warnings}."""
    # screener's own peer list is fetched only when the user asks for it
    d = fetch_company(ticker, with_peers=(peer_mode == "screener"), max_peers=max_peers)
    if not d:
        return {"error": f"Could not find '{ticker}' on screener.in. "
                         "Check the symbol (e.g. RELIANCE, TCS, HDFCBANK, BAJAJ-AUTO)."}

    h = d["history"]
    top = d["top"]
    price = top.get("Current Price")
    mcap = top.get("Market Cap")
    shares = None
    if h.get("equity_capital") and top.get("Face Value"):
        shares = h["equity_capital"] / top["Face Value"]          # Rs crore / Rs -> crore shares
    elif price and mcap:
        shares = mcap / price

    beta = compute_beta(ticker)
    yfx = yfinance_extras(ticker)

    prov: dict[str, dict] = {}

    def add(field, value, source, note=""):
        prov[field] = {"value": value, "source": source, "note": note}

    add("revenue", h.get("revenue"), "screener.in", f"{h.get('latest_period')} Sales")
    add("ebitda", h.get("ebitda"), "screener.in", f"{h.get('latest_period')} Operating Profit")
    add("ebit", h.get("ebit"), "computed", "EBITDA - Depreciation")
    add("depreciation", h.get("depreciation"), "screener.in", "P&L Depreciation")
    add("net_profit", h.get("net_profit"), "screener.in", "PAT")
    add("eps", h.get("eps"), "screener.in", "EPS in Rs")
    add("debt", h.get("debt"), "screener.in", f"Balance sheet {h.get('latest_bs_period')}")
    add("cash", h.get("cash"), "screener.in", "Investments (+ other assets if absent)")
    add("net_worth", h.get("net_worth"), "screener.in", "Equity capital + reserves")
    add("minority_interest", h.get("minority_interest"), "screener.in", "")
    add("cfo", h.get("cfo"), "screener.in", "Cash from operating activity")
    add("capex", h.get("capex_proxy"), "screener.in", "3y average of |investing cash flow|")
    add("tax_rate_pct", h.get("tax_pct"), "screener.in", "Reported effective tax %")
    add("price", price, "screener.in", "Current Price")
    add("market_cap", mcap, "screener.in", "Market Cap (Rs Cr)")
    add("shares", shares, "computed",
        "Equity capital / face value" if h.get("equity_capital") else "Market cap / price")
    add("beta", beta if beta is not None else yfx.get("yahoo_beta"),
        "computed" if beta is not None else ("yfinance" if yfx.get("yahoo_beta") else "assumed"),
        "5y weekly OLS beta vs Nifty" if beta is not None else
        ("Yahoo's beta field" if yfx.get("yahoo_beta") else "no price history available - set manually"))
    add("dividend_payout_pct", yfx.get("payout_pct"), "yfinance",
        "dividend rate / trailing EPS")
    add("dividend_yield_pct", top.get("Dividend Yield") or yfx.get("dividend_yield_pct"),
        "screener.in", "")
    add("pe", top.get("Stock P/E"), "screener.in", "")
    add("roce", top.get("ROCE"), "screener.in", "")

    # ---- peers ----
    # screener.in's peer-comparison API is unreliable for logged-out requests (it
    # returned liquor stocks for Asian Paints and holding companies for Reliance),
    # so the curated sector set is the default and screener's list is opt-in.
    curated = curated_peers(d.get("sector") or {}, ticker)
    if peer_mode == "screener" and d["peers"]["peers"]:
        peers = list(d["peers"]["peers"])
        peer_src = "screener.in industry peers"
    elif peer_mode == "screener":
        peers, peer_src = [], "screener.in returned no peers"
    elif curated:
        peers = curated
        peer_src = ("curated sector set" if d["peers"]["peers"]
                    else "curated sector set (screener.in returned no peers)")
    else:
        peers = list(d["peers"]["peers"])
        peer_src = "screener.in industry peers (no curated set for this sector)"
    peers = peers[:max(max_peers, 8)]
    if custom_peers:
        got = []
        for t in custom_peers:
            t = (t or "").strip().upper()
            if not t or t == ticker.upper():
                continue
            e = enrich_peer(t)
            if not e:
                continue
            got.append({"ticker": t, "name": t, "cmp": e.get("cmp"),
                        "mar_cap": e.get("mar_cap"), "pe": e.get("pe"),
                        "div_yld": e.get("div_yld"), "roce": e.get("roce"),
                        "roe": e.get("roe"), "book_value": e.get("book_value"),
                        "debt": e.get("debt"), "cash": e.get("cash"),
                        "ebitda": e.get("ebitda"), "eps": e.get("eps"),
                        "source": "user"})
        peers = got or peers
        peer_src = "user-specified tickers"

    inputs = {
        "price": price,
        "market_cap": mcap,
        "shares": round(shares, 4) if shares else None,
        "beta": beta if beta is not None else (yfx.get("yahoo_beta") or 1.0),
        "pe": top.get("Stock P/E"),
        "dividend_yield_pct": top.get("Dividend Yield") or yfx.get("dividend_yield_pct"),
        "roce": top.get("ROCE"),
        "dividend_payout_pct": yfx.get("payout_pct") or h.get("dividend_payout_pct"),
        "nwc_pct_revenue": nwc_pct(h),
    }

    return {
        "company": {
            "ticker": d["ticker"], "name": d["name"], "url": d["url"],
            "sector": d.get("sector") or {}, "top": top,
            "history": h, "tables": d["tables"],
        },
        "inputs": inputs,
        "provenance": prov,
        "peers": peers,
        "peer_source": peer_src,
        "peer_median_reported": d["peers"].get("median"),
        "warnings": [],
    }


def nwc_pct(h: dict) -> Optional[float]:
    """Working capital as % of revenue, from screener's cash-conversion cycle.

    NWC/revenue = CCC_days / 365. Two guardrails matter more than precision here:
    a negative cycle (FMCG, retail funding growth from suppliers) must not be allowed
    to manufacture free cash flow, and a very long cycle -- TITAN's gold inventory runs
    to ~210 days -- must not be allowed to consume all of it. Both are clamped; the
    field is editable in Assumptions.
    """
    days = h.get("ccc_days")
    if days is None or not h.get("revenue"):
        return None
    pct = max(-5.0, min(20.0, days / 365.0 * 100.0))
    return round(pct, 2)


# --------------------------------------------------------------------------- #
_PEER_MAP = None


def _peer_map() -> dict:
    global _PEER_MAP
    if _PEER_MAP is None:
        import json
        import os
        pth = os.path.join(os.path.dirname(__file__), "sector_peers.json")
        with open(pth) as f:
            _PEER_MAP = json.load(f)["map"]
    return _PEER_MAP


def curated_peers(sector: dict, our_ticker: str) -> list[dict]:
    """Size-similar curated comparables, enriched from screener.in."""
    m = _peer_map()
    tickers: list[str] = []
    for key in ("Industry", "Broad Industry", "Sector", "Broad Sector"):
        k = (sector.get(key) or "").strip()
        if k in m:
            tickers = m[k]
            break
    if not tickers:
        return []
    tickers = [t for t in tickers if t.upper() != our_ticker.upper()][:8]
    from concurrent.futures import ThreadPoolExecutor
    out = []
    with ThreadPoolExecutor(max_workers=min(3, len(tickers))) as ex:
        res = list(ex.map(enrich_peer, tickers))
    for t, e in zip(tickers, res):
        if not e:
            continue
        out.append({"ticker": t, "name": t, "cmp": e.get("cmp"), "mar_cap": e.get("mar_cap"),
                    "pe": e.get("pe"), "div_yld": e.get("div_yld"), "roce": e.get("roce"),
                    "roe": e.get("roe"), "book_value": e.get("book_value"),
                    "debt": e.get("debt"), "cash": e.get("cash"), "ebitda": e.get("ebitda"),
                    "eps": e.get("eps"), "source": "curated"})
    return out


# --------------------------------------------------------------------------- #
def from_screener_excel(raw: bytes) -> dict:
    """
    Parse a screener.in Excel export (the file behind its 'Export' button, which is
    login-gated -- so users download it while signed in and upload it here), or
    pasted tab-separated text.
    """
    try:
        sheets = pd.read_excel(io.BytesIO(raw), sheet_name=None, header=None)
        raw_table = list(sheets.values())[0]
    except Exception:
        text = raw.decode("utf-8", errors="replace")
        rows = [ln.split("\t") if "\t" in ln else ln.split()
                for ln in text.splitlines() if ln.strip()]
        if len(rows) < 3:
            return {"error": "Could not read that file. Upload the .xls/.xlsx you got from "
                             "screener.in's Export button, or paste the tables as "
                             "tab-separated text."}
        w = max(len(r) for r in rows)
        raw_table = pd.DataFrame([r + [""] * (w - len(r)) for r in rows])
    return _normalise_export(raw_table, source="screener.in export / paste")


SECTION_TITLES = {
    "consolidated profit & loss", "consolidated balance sheet", "consolidated cash flow",
    "profit & loss", "balance sheet", "cash flow", "ratios", "data",
}

SECTION_HINTS = {
    "profit-loss": ["Sales", "Revenue", "Operating Profit", "Financing Profit", "Net Profit"],
    "balance-sheet": ["Equity Capital", "Reserves", "Borrowings", "Borrowing", "Other Liabilities"],
    "cash-flow": ["Cash from Operating Activity", "Cash from Investing Activity"],
}


def _normalise_export(df: pd.DataFrame, source: str) -> dict:
    """Rebuild screener-shaped frames from an export.

    Three layouts occur in practice:
      A. our own .xlsx           ["Sales", 899041, 962820, ...]        header row above
      B. screener's text export  ["", "Sales", "Mar 2024", ...]        spacer column
      C. flat paste              ["Sales", "Mar 2024", ...]            no header row
    Normalising the label column first makes one rule work for all three.
    """
    lines = []
    for _, row in df.iterrows():
        cells = []
        for v in row.tolist():
            if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
                cells.append("")
            elif isinstance(v, float):
                cells.append(str(int(v)) if v == int(v) else repr(v))
            else:
                cells.append(str(v).strip())
        lines.append(cells)
    lines = [ln for ln in lines if any(c for c in ln)]
    if len(lines) < 3:
        return {"error": "Nothing readable in that upload."}

    def nums_of(cells):
        return [_num(c) for c in cells]

    # ---- layout B: a spacer column in front of the labels ---------------------
    nonblank = [ln for ln in lines if any(ln)]
    col0 = sum(1 for ln in nonblank if len(ln) > 0 and ln[0].strip())
    col1 = sum(1 for ln in nonblank if len(ln) > 1 and ln[1].strip())
    if col1 >= 3 and col0 <= max(2, col1 // 4):
        lines = [ln[1:] for ln in lines]

    starts = {h.lower(): sec for sec, hints in SECTION_HINTS.items() for h in hints[:1]}

    def mostly_text(cells):
        """True when the row holds period labels ('Mar 2024'), not figures.
        A row of empty strings is not a header -- it is just padding."""
        cells = [c for c in cells if c != ""]
        if not cells:
            return False
        labels = [c for c in cells
                  if _num(c) is None and c.strip().lower() not in SECTION_TITLES]
        return len(labels) >= max(1, len(cells) // 2)

    section_rows: dict[str, list[list[str]]] = {k: [] for k in SECTION_HINTS}
    columns: dict[str, list[str]] = {k: [] for k in SECTION_HINTS}
    cur = None
    for i, ln in enumerate(lines):
        first = ln[0].strip()
        if not first or first.lower() in SECTION_TITLES:
            continue
        sec = starts.get(first.lower())
        if sec and not columns[sec]:
            cur = sec
            nxt = lines[i + 1][1:] if i + 1 < len(lines) else []
            if i > 0 and mostly_text(lines[i - 1][1:]) \
                    and lines[i - 1][0].strip().lower() not in SECTION_TITLES:
                columns[sec] = [c for c in lines[i - 1][1:] if c]      # layout A
            elif nxt and mostly_text(nxt):
                hdr = [c for c in nxt if c]                            # layout B
                if hdr and hdr[0].strip().lower() == first.lower():
                    hdr = hdr[1:]        # the anchor word repeated in the header row
                columns[sec] = hdr
            elif mostly_text(ln[1:]):
                columns[sec] = [c for c in ln[1:] if c]                # layout C
                continue                                               # that row was the header
        elif sec:
            cur = sec
        if cur is None:
            continue
        section_rows[cur].append(ln)

    def frame(sec):
        rows = section_rows.get(sec) or []
        if not rows:
            return None
        data: dict[str, list] = {}
        for r in rows:
            lab = r[0].replace("+", "").strip()
            if not lab or lab in data:
                continue
            data[lab] = [_num(v) for v in r[1:]]
        if not data:
            return None
        width = max(len(v) for v in data.values())
        cols = [c for c in (columns.get(sec) or []) if c]
        cols = (cols + [f"Y{i + 1}" for i in range(len(cols), width)])[:width]
        out = pd.DataFrame.from_dict(
            {k: (v + [None] * (width - len(v)))[:width] for k, v in data.items()},
            orient="index")
        out.columns = cols
        return out

    pnl, bs, cf = frame("profit-loss"), frame("balance-sheet"), frame("cash-flow")
    if pnl is None:
        known = {r[0].strip().lower() for sec in ("balance-sheet", "cash-flow")
                 for r in section_rows[sec]}
        fallback = [ln for ln in lines
                    if ln[0].strip() and ln[0].strip().lower() not in known
                    and ln[0].strip().lower() not in SECTION_TITLES
                    and any(_num(c) is not None for c in ln[1:])]
        if fallback:
            section_rows["profit-loss"] = fallback
            if not columns["profit-loss"]:
                columns["profit-loss"] = [f"Y{i + 1}"
                                          for i in range(max(len(ln) for ln in fallback) - 1)]
            pnl = frame("profit-loss")
    if pnl is None:
        return {"error": "No profit & loss section found in that export. Expected a row "
                         "starting with 'Sales' or 'Revenue'."}
    hist = build_history(pnl, bs, cf)

    price = shares = None
    for ln in lines[:40]:
        low = [c.lower() for c in ln]
        if price is None and any("current price" in c for c in low):
            for c in ln[1:]:
                v = _num(c)
                if v:
                    price = v
                    break
        if shares is None and any("shares" in c for c in low):
            for c in ln[1:]:
                v = _num(c)
                if v:
                    shares = v
                    break
    if shares and shares > 1e7:
        shares = shares / 1e7           # absolute shares -> crore

    return {
        "company": {"ticker": "", "name": "Imported company", "url": None, "sector": {},
                    "top": {}, "history": hist,
                    "tables": {"pnl": frame_out(pnl), "bs": frame_out(bs),
                               "cf": frame_out(cf)}},
        "inputs": {"price": price, "market_cap": None, "shares": shares, "beta": 1.0,
                   "nwc_pct_revenue": nwc_pct(hist)},
        "provenance": {}, "peers": [], "peer_source": source,
        "peer_median_reported": None, "warnings": [], "imported": True,
    }

# ============== app ==============
"""
DCF & Relative Valuation Studio -- Flask app.

Data: screener.in (financials, ratios, peers) + computed beta vs Nifty.
Model: engine/valuation.py
"""




app = Flask(__name__, template_folder=".")
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024

# On a host that recycles the container (Render/Railway/Fly free tiers) an in-memory
# cache is lost on every restart, so parsed companies are mirrored to disk.
CACHE_DIR = os.environ.get("CACHE_DIR", os.path.join(os.path.dirname(__file__), "cache"))
TTL = 15 * 60
_CACHE: dict[str, tuple[float, dict]] = {}


def _cache_get(key: str):
    hit = _CACHE.get(key)
    if hit and time.time() - hit[0] < TTL:
        return hit[1]
    try:
        path = os.path.join(CACHE_DIR, key.replace("|", "_")[:120] + ".json")
        if os.path.exists(path) and time.time() - os.path.getmtime(path) < 3600:
            with open(path) as f:
                d = json.load(f)
            _CACHE[key] = (time.time(), d)
            return d
    except Exception:
        pass
    return None


def _cache_put(key: str, value: dict):
    _CACHE[key] = (time.time(), value)
    try:
        os.makedirs(CACHE_DIR, exist_ok=True)
        with open(os.path.join(CACHE_DIR, key.replace("|", "_")[:120] + ".json"), "w") as f:
            json.dump(value, f)
    except Exception:
        pass


def _clean(x, cast=float):
    if x is None or x == "":
        return None
    try:
        return cast(x)
    except (TypeError, ValueError):
        return None


def _params(body: dict, inputs: dict) -> dict:
    """Precedence: user override (request body) -> fetched inputs -> engine default."""
    p = dict(DEFAULTS)

    def val(key, default=None, cast=float):
        b = body.get(key)
        if b not in (None, ""):
            return cast(b)
        v = (inputs or {}).get(key)
        if v not in (None, ""):
            return v
        return default

    p["risk_free_pct"] = val("risk_free_pct", DEFAULTS["risk_free_pct"])
    p["erp_pct"] = val("erp_pct", DEFAULTS["erp_pct"])
    p["crp_pct"] = val("crp_pct", 0.0)
    p["beta"] = val("beta", 1.0)
    p["target_debt_pct"] = val("target_debt_pct", None)
    p["cost_of_debt_pct"] = val("cost_of_debt_pct", None)
    p["tax_rate_pct"] = val("tax_rate_pct", None)
    p["years"] = max(3, min(20, int(val("years", 10, int))))
    p["revenue_growth_pct"] = val("revenue_growth_pct", None)
    p["growth_fade_pct"] = val("growth_fade_pct", DEFAULTS["growth_fade_pct"])
    p["ebit_margin_pct"] = val("ebit_margin_pct", None)
    p["target_margin_pct"] = val("target_margin_pct", None)
    p["margin_fade_years"] = int(val("margin_fade_years", 5, int))
    p["capex_pct_revenue"] = val("capex_pct_revenue", None)
    p["dep_pct_revenue"] = val("dep_pct_revenue", None)
    p["nwc_pct_revenue"] = val("nwc_pct_revenue", 0.0) or 0.0
    p["terminal_growth_pct"] = val("terminal_growth_pct", 6.0)
    p["terminal_method"] = body.get("terminal_method") or inputs.get("terminal_method") or "gordon"
    p["exit_ev_ebitda"] = val("exit_ev_ebitda", None)
    p["discount_convention"] = body.get("discount_convention") or inputs.get("discount_convention") or "end"
    p["payout_pct"] = val("payout_pct", None)
    p["target_payout_pct"] = val("target_payout_pct", None)
    p["ddm_growth_pct"] = val("ddm_growth_pct", None)
    p["peer_bands_pct"] = val("peer_bands_pct", 25.0)
    p["dividend_payout_pct"] = val("dividend_payout_pct", None)

    p["price"] = val("price", None)
    p["market_cap"] = val("market_cap", None)
    p["shares"] = val("shares", None)
    p["pe"] = val("pe", None)
    p["roce"] = val("roce", None)
    p["dividend_yield_pct"] = val("dividend_yield_pct", None)
    return p


def _round(o, nd=4):
    if isinstance(o, dict):
        return {k: _round(v, nd) for k, v in o.items()}
    if isinstance(o, list):
        return [_round(v, nd) for v in o]
    if isinstance(o, float):
        return None if (math.isnan(o) or math.isinf(o)) else round(o, nd)
    return o


def _model(company: dict, inputs: dict, provenance: dict, peers: list,
           peer_source: str, body: dict) -> dict:
    hist = company["history"]
    p = _params(body, inputs)
    cc = cost_of_capital(p, hist)
    fin = is_financial(company.get("sector") or {}, hist)

    dcf = dcf_error = None
    try:
        dcf = run_dcf(p, hist, cc)
    except Exception as e:
        dcf_error = str(e)

    shell = {"company": company, "params": p, "wacc": cc, "is_financial": fin,
             "provenance": provenance, "peers": peers, "peer_source": peer_source,
             "fetched_at": time.time()}

    if dcf is None:
        rim = residual_income(p, hist, cc) if fin else None
        if rim and rim.get("value_per_share") is not None:
            # a bank's DCF is expected to be meaningless -- the RIM carries the answer
            v = rim["value_per_share"]
            shell.update({
                "dcf": None, "dcf_error": dcf_error, "ddm": None, "relative": None,
                "sensitivity": None, "scenarios": None, "implied": None,
                "residual_income": rim,
                "football": [{"method": "Residual income (financials)", "low": v * 0.8,
                              "mid": v, "high": v * 1.2}],
                "warnings": warnings(p, hist, cc, None, None, financial=True),
            })
            return shell
        shell.update({"error": f"DCF failed: {dcf_error}", "dcf": None,
                      "residual_income": rim, "football": [],
                      "warnings": warnings(p, hist, cc, None, None, financial=fin)})
        return shell

    dd = ddm(p, hist, cc)
    rel = relative(p, hist, cc, peers, dcf)
    sens = sensitivity(p, hist, cc, dcf)
    scens = scenarios(p, hist, cc)

    rim = residual_income(p, hist, cc) if fin else None
    ff = football_field(dcf, dd, rel, sens, scens)
    if rim and rim.get("value_per_share"):
        v = rim["value_per_share"]
        ff.insert(0, {"method": "Residual income (financials)", "low": v * 0.8,
                      "mid": v, "high": v * 1.2})

    shell.update({
        "dcf": dcf, "dcf_error": dcf_error, "ddm": dd, "relative": rel,
        "sensitivity": sens, "scenarios": scens,
        "implied": implied_expectations(p, hist, cc, dcf),
        "football": ff, "residual_income": rim,
        "warnings": warnings(p, hist, cc, dcf, rel, financial=fin),
    })
    return shell


def _cached(ticker: str, custom_peers, peer_mode: str):
    key = f"{ticker}|{','.join(custom_peers or [])}|{peer_mode}"
    hit = _cache_get(key)
    if hit:
        return hit, True
    d = build(ticker, custom_peers=custom_peers,
                       peer_mode=peer_mode)
    if not d.get("error"):
        _cache_put(key, d)
    return d, False


# --------------------------------------------------------------------------- #
@app.get("/")
def index():
    return render_template("index.html")


@app.post("/api/analyze")
def analyze():
    body = request.get_json(force=True, silent=True) or {}
    ticker = (body.get("ticker") or "").strip().upper()
    custom = body.get("custom_peers") or []
    if isinstance(custom, str):
        custom = [t.strip() for t in custom.replace(";", ",").split(",")]
    custom = [t for t in custom if t]
    peer_mode = body.get("peer_mode") or ("screener" if body.get("use_screener_peers") else "curated")

    if not ticker:
        return jsonify({"error": "Enter a screener.in ticker (e.g. RELIANCE, TCS, HDFCBANK)."}), 400

    d, from_cache = _cached(ticker, custom, peer_mode)
    if d.get("error"):
        return jsonify(d), 404

    out = _model(d["company"], d["inputs"], d["provenance"], d["peers"],
                 d["peer_source"], body)
    out["from_cache"] = from_cache
    return jsonify(_round(out))


@app.post("/api/recalc")
def recalc():
    """Re-run the model on a payload already returned by /api/analyze, with new
    assumptions -- no re-fetch."""
    body = request.get_json(force=True, silent=True) or {}
    base = body.get("base") or {}
    company = base.get("company") or {}
    if not company.get("history"):
        return jsonify({"error": "Missing company payload. Analyse a ticker first."}), 400
    out = _model(company, base.get("inputs") or {}, base.get("provenance") or {},
                 base.get("peers") or [], base.get("peer_source") or "", body)
    return jsonify(_round(out))


@app.post("/api/import")
def import_excel():
    """Upload a screener.in Excel export (or paste its text) -> parsed company."""
    raw = None
    if "file" in request.files:
        raw = request.files["file"].read()
    elif request.data:
        raw = request.data
    if not raw:
        return jsonify({"error": "Attach the screener.in .xls/.xlsx file."}), 400
    d = from_screener_excel(raw)
    if d.get("error"):
        return jsonify(d), 400
    # an export carries no share count, so let the caller supply one
    inp = d["inputs"]
    if not inp.get("shares"):
        inp["shares"] = _clean(request.args.get("shares"))
    if not inp.get("price"):
        inp["price"] = _clean(request.args.get("price"))
    if not inp.get("shares"):
        return jsonify({"error": "That file has no share count, so no per-share value can be "
                                 "produced. Re-upload with ?shares=<crore shares>&price=<CMP>, "
                                 "or use the manual entry form.",
                        "parsed": _round({"company": d["company"]})}), 400
    out = _model(d["company"], inp, {}, d["peers"], d["peer_source"], {})
    return jsonify(_round(out))


@app.post("/api/peers")
def peers_only():
    """Refresh the comp set: user tickers, screener industry peers, or curated."""
    body = request.get_json(force=True, silent=True) or {}
    ticker = (body.get("ticker") or "").strip().upper()
    custom = body.get("custom_peers") or []
    if isinstance(custom, str):
        custom = [t.strip() for t in custom.replace(";", ",").split(",")]
    custom = [t for t in custom if t]
    mode = body.get("mode") or "screener"

    if custom:
        peers, src = [], "user-specified tickers"
        for t in custom:
            e = enrich_peer(t.strip().upper())
            if e:
                peers.append({"ticker": t.strip().upper(), "name": t.strip().upper(),
                              "cmp": e.get("cmp"), "mar_cap": e.get("mar_cap"),
                              "pe": e.get("pe"), "div_yld": e.get("div_yld"),
                              "roce": e.get("roce"), "roe": e.get("roe"),
                              "book_value": e.get("book_value"), "debt": e.get("debt"),
                              "cash": e.get("cash"), "ebitda": e.get("ebitda"),
                              "eps": e.get("eps"), "source": "user"})
    elif mode == "curated" and ticker:
        d = build(ticker, peer_mode="none")
        if d.get("error"):
            return jsonify(d), 404
        peers = curated_peers(d["company"].get("sector") or {}, ticker)
        src = "curated sector set"
    else:
        wid = warehouse_id_for(ticker)
        if not wid:
            return jsonify({"error": f"Unknown ticker '{ticker}'."}), 404
        pr = fetch_peers(wid, f"https://www.screener.in/company/{ticker}/consolidated/")
        peers, src = pr["peers"][:8], "screener.in industry peers"
    return jsonify(_round({"peers": peers, "peer_source": src}))


@app.post("/api/export")
def export_xlsx():
    body = request.get_json(force=True, silent=True) or {}
    company = body.get("company") or {}
    tables = company.get("tables") or {}
    if not tables.get("pnl"):
        return jsonify({"error": "No parsed statements to export."}), 400
    raw = build(tables, company.get("name") or "Company", company.get("ticker") or "")
    import io
    fn = f"{(company.get('ticker') or 'company')}-screener-export.xlsx"
    return send_file(io.BytesIO(raw), as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/health")
def health():
    """Liveness probe for Render/Railway/Fly."""
    return jsonify({"ok": True, "cache": len(_CACHE), "ts": time.time()})


@app.errorhandler(Exception)
def _on_error(e):
    """Never leak a stack trace into the browser -- the UI expects JSON."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return jsonify({"error": e.description}), e.code
    app.logger.exception("unhandled")
    return jsonify({"error": f"Unexpected server error: {type(e).__name__}: {e}"}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8000)), debug=False, threaded=True)
